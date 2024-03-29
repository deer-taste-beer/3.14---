import pandas as pd
import openpyxl
from scipy import interpolate
from SALib.sample import saltelli
from SALib.analyze import sobol
import numpy as np
import math
# 엑셀 파일 경로
excel_path = r'C:\Users\poip8\Desktop\drive-download-20240313T030339Z-001 (2)\GT2.xlsx'

# 엑셀 파일 읽기
df = pd.read_excel(excel_path)
wb = openpyxl.load_workbook(excel_path)
ws = wb.active

# GSA 분석을 위한 열 범위 설정 (B열부터 T열까지)
#columns_to_analyze = ['B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']
# GSA 분석을 위한 열 범위 설정 (B열부터 T열까지)
columns_to_analyze = [str(ws.cell(row=1, column=j).value) for j in range(2, 21) if j != 4]

# Fix할 열(E열)의 데이터 가져오기
data_Fix = [cell.value for cell in ws['E'][19:432015]]  # E열 데이터를 Fix 열로 사용

# 각 열에서 이상치와 결측치 제거 후 선형 보간
for col in columns_to_analyze:
    # 이상치 제거 (예: 3시그마 이상의 값 제거)
    sigma = df.loc[19:432014, col].std()
    
    df.loc[19:432014, col] = df.loc[19:432014, col].mask((df.loc[19:432014, col] - df.loc[19:432014, col].mean()).abs() > 3 * sigma)
    
    # 결측치 보간
    df.loc[19:432014, col] = df.loc[19:432014, col].interpolate(method='linear')
    
# GSA 분석을 위한 데이터 프레임 생성
data_to_analyze = {}
for col in columns_to_analyze:
    data_to_analyze[col] = df.loc[19:432014, col].tolist()

# 데이터 프레임으로 변환
df_to_analyze = pd.DataFrame(data_to_analyze)

# Sobol 분석을 위한 입력 변수 설정
problem = {
    'num_vars': len(columns_to_analyze),
    'names': columns_to_analyze,
    'bounds': [[min(df_to_analyze[col]), max(df_to_analyze[col])] for col in columns_to_analyze]
}

# 분석을 위한 샘플 생성
param_values = sobol.sample(problem, 1000)
base_sample_size = 1000
# 분석을 위한 샘플 수 계산
num_vars = len(problem['names'])
total_sample_size = base_sample_size * (2 * num_vars + 2)

# 모델 실행 및 분석
Y = df_to_analyze.values
sobol_indices = sobol.analyze(problem, Y, print_to_console=False)

# 결과 출력
print("Sobol First Order Indices:")
for i, col in enumerate(columns_to_analyze):
    print(col, sobol_indices['S1'][i])

print("\nSobol Total Order Indices:")
for i, col in enumerate(columns_to_analyze):
    print(col, sobol_indices['ST'][i])
